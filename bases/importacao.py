"""Importa/exporta a base de atletas via planilha (.xlsx).

Fluxo de enriquecimento: exportar → completar dados no Excel → reimportar. O
upsert casa cada linha por **Player ID** (quando há) ou pelo **nome
normalizado**, então reimportar não duplica ninguém. Na atualização, só
sobrescreve com valores preenchidos (célula vazia não apaga dado existente)."""

import re
import unicodedata
from datetime import date, datetime
from io import BytesIO

import openpyxl
from django.db import transaction

from .models import Atleta, Participacao

# Ordem canônica das colunas (usada no export; o import aceita variações).
COLUNAS = [
    ("Nome", "nome"),
    ("Player ID", "player_id"),
    ("Gênero", "genero"),
    ("Data de nascimento", "data_nascimento"),
    ("Classe", "classe"),
    ("Treinador", "treinador"),
    ("Telefone", "telefone"),
    ("E-mail", "email"),
    ("Região", "regiao"),
    ("Filiação", "filiacao"),
    ("Local 1", "local1"),
    ("Local 2", "local2"),
    ("Local 3", "local3"),
]

_HEADER = {
    "NOME": "nome", "PLAYER ID": "player_id", "ID": "player_id",
    "GENERO": "genero", "SEXO": "genero",
    "DATA DE NASCIMENTO": "data_nascimento", "NASCIMENTO": "data_nascimento",
    "CLASSE": "classe", "TREINADOR": "treinador", "PROFESSOR": "treinador",
    "TELEFONE": "telefone", "CELULAR": "telefone",
    "E-MAIL": "email", "EMAIL": "email",
    "REGIAO": "regiao", "FILIACAO": "filiacao", "FILIACAO ": "filiacao",
    "LOCAL 1": "local1", "LOCAL 2": "local2", "LOCAL 3": "local3",
    "CLUBE": "local1", "LOCAL": "local1",
}

_TEXTO = {"nome", "player_id", "genero", "classe", "treinador",
          "telefone", "email", "regiao", "filiacao", "local1", "local2", "local3"}


def _norm(texto):
    txt = unicodedata.normalize("NFKD", str(texto or ""))
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", txt).strip().upper()


def _parse_data(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    return None


def _texto(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def importar_atletas(caminho):
    """Lê o .xlsx e faz upsert. Devolve (criados, atualizados)."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not linhas:
        return 0, 0

    cabecalho = list(linhas[0])
    mapa = {}
    for i, h in enumerate(cabecalho):
        campo = _HEADER.get(_norm(h))
        if campo and campo not in mapa:
            mapa[campo] = i
    if "nome" not in mapa:
        raise ValueError("A planilha precisa de uma coluna 'Nome'.")

    # pré-carrega existentes para casar sem uma query por linha
    por_pid, por_nome = {}, {}
    for a in Atleta.objects.all():
        if a.player_id:
            por_pid[a.player_id] = a
        por_nome.setdefault(_norm(a.nome), a)

    criados = atualizados = 0
    novos = []
    with transaction.atomic():
        for row in linhas[1:]:
            if row is None:
                continue

            def val(campo):
                i = mapa.get(campo)
                return row[i] if (i is not None and i < len(row)) else None

            nome = _texto(val("nome"))
            if not nome:
                continue
            pid = _texto(val("player_id"))

            obj = (por_pid.get(pid) if pid else None) or por_nome.get(_norm(nome))
            novo = obj is None
            if novo:
                obj = Atleta(nome=nome)

            # atualiza só o que veio preenchido (enriquecimento não apaga dado)
            for campo in mapa:
                if campo == "data_nascimento":
                    d = _parse_data(val(campo))
                    if d:
                        obj.data_nascimento = d
                else:
                    v = _texto(val(campo))
                    if v:
                        setattr(obj, campo, v)
            obj.nome = nome
            obj.save()

            if novo:
                criados += 1
                if obj.player_id:
                    por_pid[obj.player_id] = obj
                por_nome.setdefault(_norm(nome), obj)
            else:
                atualizados += 1
    return criados, atualizados


def exportar_atletas():
    """Gera o .xlsx da base inteira (bytes), pronto para enriquecer e reimportar."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atletas"
    ws.append([rot for rot, _ in COLUNAS])
    for a in Atleta.objects.all():
        linha = []
        for _, campo in COLUNAS:
            if campo == "data_nascimento":
                linha.append(a.data_nascimento.strftime("%Y-%m-%d") if a.data_nascimento else "")
            else:
                linha.append(getattr(a, campo) or "")
        ws.append(linha)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================ Participações ============================

_COLS_PART = [
    ("Campeonato", "campeonato"), ("Data", "data"), ("Categoria Disputada", "categoria"),
    ("Atleta", "atleta_nome"), ("Colocação", "colocacao"), ("Pontuação", "pontuacao"),
    ("Academia e Clube", "academia"), ("Treinador", "treinador"),
]

_HEADER_PART = {
    "CAMPEONATO": "campeonato", "EVENTO": "campeonato",
    "DATA": "data",
    "CATEGORIA DISPUTADA": "categoria", "CATEGORIA": "categoria",
    "ATLETA": "atleta_nome", "NOME": "atleta_nome", "JOGADOR": "atleta_nome",
    "COLOCACAO": "colocacao", "RESULTADO": "colocacao", "POSICAO": "colocacao",
    "PONTUACAO": "pontuacao", "PONTOS": "pontuacao",
    "ACADEMIA E CLUBE": "academia", "ACADEMIA/CLUBE": "academia",
    "CLUBE": "academia", "ACADEMIA": "academia",
    "TREINADOR": "treinador", "PROFESSOR": "treinador",
}


def _int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def importar_participacoes(caminho):
    """Lê o .xlsx de resultados e substitui a base de participações, linkando cada
    linha ao atleta pelo nome. Devolve (total, casados, sem_atleta)."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not linhas:
        return 0, 0, 0

    cabecalho = list(linhas[0])
    mapa = {}
    for i, h in enumerate(cabecalho):
        campo = _HEADER_PART.get(_norm(h))
        if campo and campo not in mapa:
            mapa[campo] = i
    if "atleta_nome" not in mapa:
        raise ValueError("A planilha precisa de uma coluna 'Atleta'.")

    por_nome = {}
    for a in Atleta.objects.all():
        por_nome.setdefault(_norm(a.nome), a)

    objs, casados = [], 0
    for row in linhas[1:]:
        if row is None:
            continue

        def val(campo):
            i = mapa.get(campo)
            return row[i] if (i is not None and i < len(row)) else None

        nome = _texto(val("atleta_nome"))
        if not nome:
            continue
        atleta = por_nome.get(_norm(nome))
        if atleta:
            casados += 1
        objs.append(Participacao(
            atleta=atleta, atleta_nome=nome,
            campeonato=_texto(val("campeonato")), data=_parse_data(val("data")),
            categoria=_texto(val("categoria")), colocacao=_texto(val("colocacao")),
            pontuacao=_int(val("pontuacao")), academia=_texto(val("academia")),
            treinador=_texto(val("treinador")),
        ))

    with transaction.atomic():
        Participacao.objects.all().delete()
        Participacao.objects.bulk_create(objs, batch_size=1000)
    return len(objs), casados, len(objs) - casados


def exportar_participacoes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append([rot for rot, _ in _COLS_PART])
    for p in Participacao.objects.select_related("atleta"):
        ws.append([
            p.campeonato,
            p.data.strftime("%Y-%m-%d") if p.data else "",
            p.categoria, p.atleta_nome, p.colocacao,
            p.pontuacao if p.pontuacao is not None else "",
            p.academia, p.treinador,
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
