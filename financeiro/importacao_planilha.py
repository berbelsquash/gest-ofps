"""Importa a planilha financeira categorizada da FPS (fonte da verdade até maio).

Lê as abas de detalhe: Eventos, Despesas gerais, Folha FPS e Receitas — que já
estão categorizadas manualmente — para o ledger LancamentoFinanceiro.
"""

from decimal import Decimal, InvalidOperation

import openpyxl

from .models import LancamentoFinanceiro

MES_NOME = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "março": 3, "abril": 4, "maio": 5,
    "junho": 6, "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10,
    "novembro": 11, "dezembro": 12,
}


def _mes(v):
    if v is None:
        return None
    if hasattr(v, "month"):
        return v.month
    return MES_NOME.get(str(v).strip().lower())


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).replace("R$", "").replace("\xa0", "").replace(".", "").replace(",", ".").strip()
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def importar_planilha(caminho, ano=2026, ate_mes=5, substituir=True):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    regs = []

    # --- Eventos: Conta, Categoria, Evento, Valores, Mês, Data, NF, Pago, Descrição ---
    for r in wb["Eventos"].iter_rows(min_row=2, values_only=True):
        conta, cat, ev, val, mes = r[0], r[1], r[2], r[3], r[4]
        m, v = _mes(mes), _num(val)
        if not m or m > ate_mes or v is None or not conta:
            continue
        tipo = "despesa" if str(conta).strip().lower() == "despesa" else "receita"
        grupo = "Torneios" if tipo == "despesa" else "Inscrições"
        regs.append(LancamentoFinanceiro(
            ano=ano, mes=m, tipo=tipo, grupo=grupo,
            categoria=(str(cat) if cat else "")[:80],
            evento=(str(ev).strip() if ev else "")[:120],
            valor=v, descricao="", origem="planilha",
        ))

    # --- Despesas gerais: Conta, Categoria, Valores, Data, Mês, Pago, Descrição ---
    for r in wb["Despesas gerais"].iter_rows(min_row=3, values_only=True):
        conta, cat, val, data, mes = r[0], r[1], r[2], r[3], r[4]
        m, v = _mes(mes or data), _num(val)
        if not m or m > ate_mes or v is None or str(conta).strip().lower() != "despesa":
            continue
        regs.append(LancamentoFinanceiro(
            ano=ano, mes=m, tipo="despesa", grupo="Despesas gerais",
            categoria=(str(cat) if cat else "")[:80],
            descricao=(str(r[6]) if len(r) > 6 and r[6] else "")[:255],
            valor=v, origem="planilha",
        ))

    # --- Folha FPS: Conta, Remunerado, Valores, Mês ---
    for r in wb["Folha FPS"].iter_rows(min_row=3, values_only=True):
        conta, nome, val, mes = r[0], r[1], r[2], r[3]
        m, v = _mes(mes), _num(val)
        if not m or m > ate_mes or v is None or str(conta).strip().lower() != "despesa":
            continue
        regs.append(LancamentoFinanceiro(
            ano=ano, mes=m, tipo="despesa", grupo="Folha",
            categoria=(str(nome) if nome else "")[:80], valor=v, origem="planilha",
        ))

    # --- Receitas: Categoria, Descrição, Mês, Pago(valor) ---
    for r in wb["Receitas"].iter_rows(min_row=2, values_only=True):
        cat, desc, mes, pago = r[0], r[1], r[2], r[3]
        m, v = _mes(mes), _num(pago)
        if not cat or not m or m > ate_mes or v is None:
            continue
        regs.append(LancamentoFinanceiro(
            ano=ano, mes=m, tipo="receita", grupo="Receitas",
            categoria=(str(cat))[:80], descricao=(str(desc) if desc else "")[:255],
            valor=v, origem="planilha",
        ))

    if substituir:
        LancamentoFinanceiro.objects.filter(origem="planilha").delete()
    LancamentoFinanceiro.objects.bulk_create(regs, batch_size=500)
    return len(regs)
